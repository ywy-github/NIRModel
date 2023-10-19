import os
import json
import numpy as np
from sklearn.cluster import KMeans
import cv2
from imutils import build_montages
import torch.nn as nn
import torchvision.models as models
from PIL import Image
from torchvision import transforms
# from torchvision.models.resnet import ResNet18_Weights
from openpyxl import load_workbook



class Net(nn.Module):
    def __init__(self):
        super(Net, self).__init__()
        # resnet18 = models.resnet18(weights=ResNet18_Weights.IMAGENET1K_V1)
        resnet18 = models.resnet18(pretrained=True)
        self.resnet = nn.Sequential(resnet18.conv1,
                                    resnet18.bn1,
                                    resnet18.relu,
                                    resnet18.maxpool,
                                    resnet18.layer1,
                                    resnet18.layer2,
                                    resnet18.layer3,
                                    resnet18.layer4)

    def forward(self, x):
        x = self.resnet(x)
        return x

if __name__ == '__main__':
    # 预训练模型，要注意的是定义好最终输出，最好是卷积层。这里的net最后一层是一个卷积层
    net = Net().eval()

    # 样本地址，读取样本名、lable、实例
    source_path = '../../../data/image_all/'
    image_path = []
    all_images = []
    data_dict = {}


    # 打开Excel文件
    workbook = load_workbook('/home/eva-01/桌面/NIR/NIR_resnet18/shiyan/label/Phase1+2_label_shuffle.xlsx')

    # 选择工作表
    sheet = workbook['Sheet1']

    # for cell in sheet['A']:
    #     print(cell.value)
    keys = []
    values = []
    keys_train = []
    values_train = []
    keys_test = []
    values_test = []
    for row in sheet.iter_rows():
        keys.append(row[0].value)
        values.append(row[1].value)

    pairs = zip(keys, values)
    labels = []
    # 使用字典推导式创建字典
    data_dict = {k: v for k, v in pairs}
    x = 1

    for image_name in list(data_dict.keys()):

        image_path.append(os.path.join(source_path, image_name))
    for path in image_path:
        path = path.replace("\\", "/")
        try:
            image = Image.open(path).convert('RGB')
        except:
            image_path.remove(path)
            continue
        labels.append(data_dict[path.split('/')[-1]])
        image = transforms.Resize([112, 112])(image)
        image = transforms.ToTensor()(image)
        image = image.unsqueeze(0)
        image = net(image)
        image = image.reshape(-1, )
        all_images.append(image.detach().numpy())
    labels = np.array(labels)
    ##### ------------------------------------------------- #####



    # 聚类，具体聚多少类要自己定，不过有这方面工作是聚成和类别数一致，那咱们这就是2类
    clt = KMeans(n_clusters=4, n_init=500)
    clt.fit(all_images)
    labelIDs = np.unique(clt.labels_)

    # 下面是聚类可视化过程
    import matplotlib.pyplot as plt
    import seaborn as sns
    import pandas as pd
    import numpy as np
    from sklearn.manifold import TSNE
    tsne = TSNE(n_components=2, random_state=0)
    X_tsne = tsne.fit_transform(all_images)
    X_tsne_data = np.vstack((X_tsne.T, clt.labels_, labels)).T
    df_tsne = pd.DataFrame(X_tsne_data, columns=['Dim1', 'Dim2', 'clt_class', 'tumor_nature'])
    df_tsne.head()
    plt.figure(figsize=(8, 8))
    sns.scatterplot(data=df_tsne.loc[(df_tsne["tumor_nature"] == 0) & (df_tsne["clt_class"] == 0)], x='Dim1', y='Dim2', marker='*', color='blue');
    sns.scatterplot(data=df_tsne.loc[(df_tsne["tumor_nature"] == 1) & (df_tsne["clt_class"] == 0)], x='Dim1', y='Dim2', marker='o', color='blue');
    sns.scatterplot(data=df_tsne.loc[(df_tsne["tumor_nature"] == 0) & (df_tsne["clt_class"] == 1)], x='Dim1', y='Dim2', marker='*', color='black');
    sns.scatterplot(data=df_tsne.loc[(df_tsne["tumor_nature"] == 1) & (df_tsne["clt_class"] == 1)], x='Dim1', y='Dim2', marker='o', color='black');
    sns.scatterplot(data=df_tsne.loc[(df_tsne["tumor_nature"] == 0) & (df_tsne["clt_class"] == 2)], x='Dim1', y='Dim2', marker='*', color='red');
    sns.scatterplot(data=df_tsne.loc[(df_tsne["tumor_nature"] == 1) & (df_tsne["clt_class"] == 2)], x='Dim1', y='Dim2', marker='o', color='red');
    sns.scatterplot(data=df_tsne.loc[(df_tsne["tumor_nature"] == 0) & (df_tsne["clt_class"] == 3)], x='Dim1', y='Dim2', marker='*', color='green');
    sns.scatterplot(data=df_tsne.loc[(df_tsne["tumor_nature"] == 1) & (df_tsne["clt_class"] == 3)], x='Dim1', y='Dim2', marker='o', color='green');
    plt.show()
    ##### ------------------------------------------------- #####

    # 不是很重要的部分
    cluster_images = {labelID: [] for labelID in labelIDs}
    mean_list = {labelID: [] for labelID in labelIDs}
    for labelID in labelIDs:
        idxs = np.where(clt.labels_ == labelID)[0]
        show_box = []

        for i in idxs:
            image = Image.open(image_path[i]).convert('RGB')
            image = np.array(image)
            image = cv2.resize(image, (96, 96))
            image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)  # Convert from RGB to BGR format for OpenCV compatibility
            show_box.append(image)
            pixel_mean = np.mean(image[np.nonzero(image)])
            # print("Cluster {}: Image {}: Pixel Mean: {}".format(labelID, os.path.basename(image_path[i]), pixel_mean))

            # Convert np.int32 keys to int
            #         cluster_images = {str(int(labelID)): [] for labelID in labelIDs}

            cluster_images[labelID].append(os.path.basename(image_path[i]))
            mean_list[labelID].append(pixel_mean)
        n_rows = int(np.ceil(len(show_box) / 5))
        montage = build_montages(show_box, (96, 96), (5, n_rows))[0]

        title = "Type {}".format(labelID)
        cv2.imwrite(title + ".jpg", montage)

    # Save the cluster image names to a JSON file
    cluster_json = {int(key): value for key, value in cluster_images.items()}
    with open("cluster_json.json", "w", encoding="utf-8") as file:
        json.dump(cluster_json, file, ensure_ascii=False)

    # 根据聚类结果生成训练集
    for row in sheet.iter_rows(min_row=1, max_row=1563):
        keys_train.append(row[0].value)
        values_train.append(row[1].value)
    pairs_train = zip(keys_train, values_train)

    # 使用字典推导式创建字典
    data_dict_train = {k: v for k, v in pairs_train}
    train_dict = json.dumps(data_dict_train)


    # for image_name in list(data_dict.keys()):
    # with open(os.path.join(source_path, r'data/ori_enhanced/train.json')) as f:
    #     train_dict = json.load(f)

    category_dict_train = {0: [], 1: [], 2: [], 3: []}
    for key in data_dict_train.keys():
        if key in cluster_images[0]:
            category_dict_train[0].append(key)
        elif key in cluster_images[1]:
            category_dict_train[1].append(key)
        elif key in cluster_images[2]:
            category_dict_train[2].append(key)
        elif key in cluster_images[3]:
            category_dict_train[3].append(key)

    with open(os.path.join(source_path, r'category_json_new/train_kind.json'), 'w') as f:
        json.dump(category_dict_train, f)

    train_dict1 = {}
    for key in category_dict_train[0]:
        train_dict1[key] = data_dict_train[key]

    train_dict2 = {}
    for key in category_dict_train[1]:
        train_dict2[key] = data_dict_train[key]

    train_dict3 = {}
    for key in category_dict_train[2]:
        train_dict3[key] = data_dict_train[key]

    train_dict4 = {}
    for key in category_dict_train[3]:
        train_dict4[key] = data_dict_train[key]


    with open(os.path.join(source_path, r'category_json_new/train_kind1.json'), 'w') as f:
        json.dump(train_dict1, f)

    with open(os.path.join(source_path, r'category_json_new/train_kind2.json'), 'w') as f:
        json.dump(train_dict2, f)

    with open(os.path.join(source_path, r'category_json_new/train_kind3.json'), 'w') as f:
        json.dump(train_dict3, f)

    with open(os.path.join(source_path, r'category_json_new/train_kind4.json'), 'w') as f:
        json.dump(train_dict4, f)



    for row in sheet.iter_rows(min_row=1564, max_row=2204):
        keys_test.append(row[0].value)
        values_test.append(row[1].value)
    pairs_test = zip(keys_test, values_test)

    # 使用字典推导式创建字典
    data_dict_test = {k: v for k, v in pairs_test}
    test_dict = json.dumps(data_dict_test)
    category_dict_test = {0: [], 1: [], 2: [], 3: []}
    for key in data_dict_test.keys():
        if key in cluster_images[0]:
            category_dict_test[0].append(key)
        elif key in cluster_images[1]:
            category_dict_test[1].append(key)
        elif key in cluster_images[2]:
            category_dict_test[2].append(key)
        elif key in cluster_images[3]:
            category_dict_test[3].append(key)

    with open(os.path.join(source_path, r'category_json_new/train_kind.json'), 'w') as f:
        json.dump(category_dict_test, f)

    test_dict1 = {}
    for key in category_dict_test[0]:
        test_dict1[key] = data_dict_test[key]

    test_dict2 = {}
    for key in category_dict_test[1]:
        test_dict2[key] = data_dict_test[key]

    test_dict3 = {}
    for key in category_dict_test[2]:
        test_dict3[key] = data_dict_test[key]

    test_dict4 = {}
    for key in category_dict_test[3]:
        test_dict4[key] = data_dict_test[key]

    with open(os.path.join(source_path, r'category_json_new/test_kind1.json'), 'w') as f:
        json.dump(test_dict1, f)

    with open(os.path.join(source_path, r'category_json_new/test_kind2.json'), 'w') as f:
        json.dump(test_dict2, f)

    with open(os.path.join(source_path, r'category_json_new/test_kind3.json'), 'w') as f:
        json.dump(test_dict3, f)

    with open(os.path.join(source_path, r'category_json_new/test_kind4.json'), 'w') as f:
        json.dump(test_dict4, f)
    ##### ------------------------------------------------- #####

    # 打印均值和标准差
    len(train_dict1),len(train_dict2),len(train_dict3),len(train_dict4)
    # 打印均值和标准差
    print("Mean: {:.2f}, {:.2f}, {:.2f}, {:.2f}".format(np.mean(mean_list[0]), np.mean(mean_list[1]), np.mean(mean_list[2]), np.mean(mean_list[3])))
    print("Standard Deviation: {:.2f}, {:.2f}, {:.2f}, {:.2f}".format(np.std(mean_list[0]), np.std(mean_list[1]), np.std(mean_list[2]), np.std(mean_list[3])))